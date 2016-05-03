#!/usr/bin/env python3
# -*- coding: utf-8 -*-
from openpyxl import Workbook
from openpyxl import load_workbook
from datetime import datetime,timedelta
#from openpyxl.compat import range
#from openpyxl.cell import get_column_letter

today = datetime.today()
yestoday = datetime.today() - timedelta(days=1)
yyestoday = datetime.today() - timedelta(days=2)

dest_filename = 'Online_Operation'+yyestoday.strftime('%m%d')+'.xlsx'
read_filename = 'Online_Operation_Day'+yestoday.strftime('%m%d')+'.xlsx'
finally_filename ='Online_Operation'+yestoday.strftime('%m%d')+'.xlsx'


cell_list1=[]
cell_list2=[]

read_wb = load_workbook(filename = read_filename)
dest_wb = load_workbook(filename = dest_filename)

read_sheet1 = read_wb[yestoday.strftime('%Y-%m-%d')+'B端基础数据']
cell_range1 = read_sheet1['a2':'j2']
for row in cell_range1:
    for cell in row:
        cell_list1.append(cell.value) 
dest_sheet1 = dest_wb['B端基础数据']
dest_sheet1.append(cell_list1)       
           
read_sheet2 = read_wb[yestoday.strftime('%Y-%m-%d')+'C端基础数据']
cell_range2 = read_sheet2['a2':'u2']
for row in cell_range2:
    for cell in row:
        cell_list2.append(cell.value)   
dest_sheet2 = dest_wb['C端基础数据']
dest_sheet2.append(cell_list2)        
           
read_sheet3 = read_wb[yestoday.strftime('%Y-%m-%d')+'B端SKU销量']
read_rows = len(read_sheet3.rows)
dest_sheet3 = dest_wb['B端SKU销量']
cell_range3 = read_sheet3['a2':'d'+str(read_rows)]
for row in cell_range3:
    cell_list3=[]
    for cell in row:          
        cell_list3.append(cell.value)
    dest_sheet3.append(cell_list3) 
        

dest_wb.save(filename = finally_filename)