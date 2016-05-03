#!/usr/bin/env python3
# -*- coding: utf-8 -*-
import os,sys
import re
from openpyxl import Workbook
from openpyxl import load_workbook
from datetime import datetime,timedelta
#from openpyxl.cell import get_column_letter

def copy_excel(read_filename):
    cell_list1=[]
    cell_list2=[]
    
    read_wb = load_workbook(filename = read_filename)    
    sheet_names=read_wb.get_sheet_names()
    read_sheet1 = read_wb[sheet_names[0]]
    cell_range1 = read_sheet1['a2':'j2']
    for row in cell_range1:
        for cell in row:
            cell_list1.append(cell.value) 
    dest_sheet1 = dest_wb['B端基础数据']
    dest_sheet1.append(cell_list1)       
               
    read_sheet2 = read_wb[sheet_names[1]]
    cell_range2 = read_sheet2['a2':'u2']
    for row in cell_range2:
        for cell in row:
            cell_list2.append(cell.value)   
    dest_sheet2 = dest_wb['C端基础数据']
    dest_sheet2.append(cell_list2)        
               
    read_sheet3 = read_wb[sheet_names[2]]
    read_rows = len(read_sheet3.rows)
    dest_sheet3 = dest_wb['B端SKU销量']
    cell_range3 = read_sheet3['a2':'d'+str(read_rows)]
    for row in cell_range3:
        cell_list3=[]
        for cell in row:          
            cell_list3.append(cell.value)
        dest_sheet3.append(cell_list3) 

        
file_list=[f for f in os.listdir() if re.match('Online_Operation_Day',f)]
dest_wb = Workbook()
ws1 = dest_wb.active
ws1.title='B端基础数据'
ws2 = dest_wb.create_sheet(title='C端基础数据')
ws2 = dest_wb.create_sheet(title='B端SKU销量')
for f in file_list:
    copy_excel(f)
dest_wb.save(filename = 'Online_Operation.xlsx')