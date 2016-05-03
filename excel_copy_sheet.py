#!/usr/bin/env python3
# -*- coding: utf-8 -*-
from __future__ import absolute_import
from openpyxl import Workbook
from openpyxl import load_workbook

finally_filename = 'aftercopy_book.xlsx'
origin_filename = 'origin_book.xlsx'
read_wb = load_workbook(filename = origin_filename)
sheet_names = read_wb.get_sheet_names()

ws1=read_wb[sheet_names[0]]
ws2 = read_wb.create_sheet(title="copy_data")

for row_idx,row in enumerate(ws1.rows):
    for col_idx, cell in enumerate(row):
        _ = ws2.cell(column=1, row=row_idx+1, value="%s" % (cell.value))

read_wb.save(filename = finally_filename)