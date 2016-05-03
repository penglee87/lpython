#读取excel使用(支持03)  
import xlrd  
#写入excel使用(支持03)  
import xlwt3  
#读取execel使用(支持07)  
from openpyxl import Workbook  
#写入excel使用(支持07)  
from openpyxl import load_workbook  
  
  
def showexcel(path):  
    workbook=xlrd.open_workbook(path)  
    sheets=workbook.sheet_names();  
    #多个sheet时，采用下面的写法打印  
    #for sname in sheets:  
        #print(sname)  
    worksheet=workbook.sheet_by_name(sheets[0])  
    #nrows=worksheet.nrows  
    #nclows=worksheet.ncols  
    for i in range(0,worksheet.nrows):  
        row=worksheet.row(i)  
  
        for j in range(0,worksheet.ncols):  
            print(worksheet.cell_value(i,j),"\t",end="")  
  
        print()  
  
  
  
def writeexcel03(path):  
  
    wb=xlwt3.Workbook()  
    sheet=wb.add_sheet("xlwt3数据测试表")  
    value = [["名称", "hadoop编程实战", "hbase编程实战", "lucene编程实战"], ["价格", "52.3", "45", "36"], ["出版社", "机械工业出版社", "人民邮电出版社", "华夏人民出版社"], ["中文版式", "中", "英", "英"]]  
    for i in range(0,4):  
        for j in range(0,len(value[i])):  
            sheet.write(i,j,value[i][j])  
    wb.save(path)  
    print("写入数据成功！")  
  
def writeexcel07(path):  
  
    wb=Workbook()  
    #sheet=wb.add_sheet("xlwt3数据测试表")  
    sheet=wb.create_sheet(0,"xlwt3数据测试表")  
  
    value = [["名称", "hadoop编程实战", "hbase编程实战", "lucene编程实战"], ["价格", "52.3", "45", "36"], ["出版社", "机械工业出版社", "人民邮电出版社", "华夏人民出版社"], ["中文版式", "中", "英", "英"]]  
    #for i in range(0,4):  
        #for j in range(0,len(value[i])):  
            #sheet.write(i,j,value[i][j])  
  
            #sheet.append(value[i])  
    sheet.cell(row = 1,column= 2).value="温度"  
    wb.save(path)  
    print("写入数据成功！")  
  
  
def read07excel(path):  
    wb2=load_workbook(path)  
    #print(wb2.get_sheet_names())  
    ws=wb2.get_sheet_by_name("详单一")  
    row=ws.get_highest_row()  
    col=ws.get_highest_column()  
    print("列数: ",ws.get_highest_column())  
    print("行数: ",ws.get_highest_row())  
  
    for i  in range(0,row):  
        for j in range(0,col):  
            print(ws.rows[i][j].value,"\t\t",end="")  
  
        print()  
  
    #print(ws.rows[0][0].value)  
    #print(ws.rows[1][0].value)  
    #print(ws.rows[0][1].value)  
  
  
  
  
  
  
#excelpath=r"D://名称.xlsx"  
#writepath=r"D://书籍明细07.xlsx"  
#writeexcel03(writepath)  
#writeexcel07(writepath)  
  
read07path="D://名称.xlsx";  
  
  
read03path=r"E:\同义词词库.xls";  
#read07excel(read07path)  
#read07excel(read03path)  
#showexcel(excelpath);  
showexcel(read03path); 